import os
import re
import uuid
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

NOTAS_PATH = "Notas"
SALIDA_EXCEL = "Salida/planilla_ingresos.xlsx"

# Frases clave para modalidad
FRASE_ALTA_POR_BAJA = "PEDIDO DE FACTIBILIDAD DE DESIGNACIÓN DE PERSONAL"
FRASE_VACANTE_NUEVA = "DESCRIPCIÓN DEL PUESTO PARA INGRESO"

# Frases clave para vacantes
FRASES_BUSQUEDA_VACANTES = [
    "CANTIDAD DE PERSONAS INCLUIDAS EN LA PROPUESTA",
    "CANTIDAD DE VACANTES A AFECTAR COMO RESPALDO PRESUPUESTARIO",
    "CANTIDAD DE VACANTES REQUERIDAS"
]

def extraer_texto_pdf(ruta_pdf):
    try:
        with fitz.open(ruta_pdf) as doc:
            texto = ""
            for page in doc:
                texto += page.get_text()
            return texto.upper()
    except Exception as e:
        print(f"❌ Error leyendo {ruta_pdf}: {e}")
        return ""

def extraer_nota_o1(texto):
    coincidencias = re.findall(r'NO-\d{4}-\d{8}-GDEBA-[A-Z0-9]+', texto)
    return coincidencias if coincidencias else ["No detectado"]

def detectar_fecha(texto):
    match = re.search(r"(lunes|martes|miércoles|jueves|viernes|sábado|domingo)\s+(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", texto, re.IGNORECASE)
    if match:
        _, num, mes, anio = match.groups()
        try:
            meses = {
                "enero":1, "febrero":2, "marzo":3, "abril":4, "mayo":5,
                "junio":6, "julio":7, "agosto":8, "septiembre":9,
                "octubre":10, "noviembre":11, "diciembre":12
            }
            fecha = datetime(int(anio), meses[mes.lower()], int(num))
            return fecha.strftime("%d/%m/%y")
        except:
            return ""
    return ""

def detectar_modalidad_excel(ruta_excel):
    try:
        xls = pd.ExcelFile(ruta_excel)
        for hoja in xls.sheet_names:
            df = xls.parse(hoja, header=None).astype(str).applymap(lambda x: x.upper())
            if df.stack().str.contains(FRASE_ALTA_POR_BAJA).any():
                return "Altas por bajas"
            elif df.stack().str.contains(FRASE_VACANTE_NUEVA).any():
                return "Vacante nueva"
    except Exception as e:
        print(f"❌ Error leyendo Excel {ruta_excel}: {e}")
    return "No detectada"

def detectar_vacantes(excel_path):
    try:
        wb = load_workbook(excel_path, data_only=True)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            for fila in ws.iter_rows(values_only=True):
                for i, celda in enumerate(fila):
                    if celda:
                        celda_str = str(celda).strip().upper()
                        for frase in FRASES_BUSQUEDA_VACANTES:
                            if frase in celda_str:
                                for siguiente in fila[i+1:]:
                                    if isinstance(siguiente, (int, float)) and siguiente > 0:
                                        return int(siguiente)
        return "No detectado"
    except Exception as e:
        print(f"Error leyendo vacantes de {excel_path}: {e}")
        return "No detectado"

def sumar_dias_habiles(fecha_str):
    try:
        fecha = datetime.strptime(fecha_str, "%d/%m/%y")
        dias_agregados = 0
        while dias_agregados < 10:
            fecha += pd.Timedelta(days=1)
            if fecha.weekday() < 5:
                dias_agregados += 1
        return fecha.strftime("%d/%m/%y")
    except:
        return ""

def cargar_o_crear_excel():
    columnas = ["Orden", "Jurisdicción", "Modalidad", "Número O1", "Fecha O1", "Vacantes", "Fecha Plazo de respuesta"]
    if os.path.exists(SALIDA_EXCEL):
        return pd.read_excel(SALIDA_EXCEL)
    else:
        return pd.DataFrame(columns=columnas)

def guardar_excel(df):
    os.makedirs(os.path.dirname(SALIDA_EXCEL), exist_ok=True)
    df.to_excel(SALIDA_EXCEL, index=False)

def procesar_notas():
    df = cargar_o_crear_excel()
    for jurisdiccion in os.listdir(NOTAS_PATH):
        ruta_jurisdiccion = os.path.join(NOTAS_PATH, jurisdiccion)
        if not os.path.isdir(ruta_jurisdiccion):
            continue
        print(f"📁 Jurisdicción: {jurisdiccion}")
        for carpeta_fecha in os.listdir(ruta_jurisdiccion):
            ruta_carpeta = os.path.join(ruta_jurisdiccion, carpeta_fecha)
            if not os.path.isdir(ruta_carpeta):
                continue
            print(f"  🔹 Nota: {carpeta_fecha}")

            pdfs = [f for f in os.listdir(ruta_carpeta) if f.lower().endswith(".pdf")]
            excels = [f for f in os.listdir(ruta_carpeta) if f.lower().endswith((".xls", ".xlsx"))]

            texto_principal = ""
            textos = []

            for pdf in pdfs:
                path_pdf = os.path.join(ruta_carpeta, pdf)
                texto = extraer_texto_pdf(path_pdf)
                textos.append(texto)
                if "no-" in pdf.lower() and ("SOLICITUD COBERTURA CARGOS" in texto or "SOLICITA LA APROBACIÓN DE" in texto):
                    texto_principal = texto

            if not texto_principal:
                texto_principal = " ".join(textos)

            numeros_o1 = extraer_nota_o1(texto_principal)
            fecha_o1 = detectar_fecha(texto_principal)
            fecha_plazo = sumar_dias_habiles(fecha_o1)

            modalidad = "No detectada"
            vacantes = "No detectado"

            for excel in excels:
                ruta_excel = os.path.join(ruta_carpeta, excel)
                modalidad = detectar_modalidad_excel(ruta_excel)
                vacantes = detectar_vacantes(ruta_excel)
                if modalidad != "No detectada" or vacantes != "No detectado":
                    break

            for numero_o1 in numeros_o1:
                if not df["Número O1"].astype(str).str.contains(numero_o1).any():
                    fila = {
                        "Orden": str(uuid.uuid4())[:8],
                        "Jurisdicción": jurisdiccion,
                        "Modalidad": modalidad,
                        "Número O1": numero_o1,
                        "Fecha O1": fecha_o1,
                        "Vacantes": vacantes,
                        "Fecha Plazo de respuesta": fecha_plazo
                    }
                    df = pd.concat([df, pd.DataFrame([fila])], ignore_index=True)
                else:
                    print(f"⚠️ Número ya registrado: {numero_o1}")

    guardar_excel(df)
    print("✅ Análisis completado.")

procesar_notas()
